#!/usr/bin/env python3
"""即梦AI脚本生成助手"""
import os
import glob
from flask import Flask, render_template_string, request, jsonify, send_file
import openpyxl

app = Flask(__name__)

HTML = '''
<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>AI脚本助手</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,sans-serif;background:#000;color:#fff;padding:20px}
.container{max-width:1200px;margin:0 auto}
.header{text-align:center;padding:30px;background:linear-gradient(135deg,#1a1a2e,#16213e);border-radius:20px;margin-bottom:20px}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}
.card{background:#1C1C1E;border-radius:16px;padding:20px;margin-bottom:20px}
.card h2{font-size:18px;margin-bottom:15px}
textarea{width:100%;background:#2C2C2E;border:none;border-radius:10px;padding:12px;color:#fff;min-height:100px;margin-bottom:10px}
.btn{background:#007AFF;color:#fff;padding:10px 20px;border:none;border-radius:8px;cursor:pointer;margin-right:10px}
.btn:hover{background:#0077ED}
.file-item{padding:10px;background:#2C2C2E;border-radius:8px;margin-bottom:8px;cursor:pointer}
.file-item:hover{background:#3A3A3C}
.chat{height:300px;overflow-y:auto;background:#2C2C2E;border-radius:10px;padding:15px;margin-bottom:10px}
.msg{margin-bottom:10px;padding:10px;border-radius:8px;max-width:80%}
.msg-ai{background:#1C1C1E}
.msg-user{background:#007AFF;margin-left:auto}
</style></head>
<body><div class="container">
<div class="header"><h1>AI脚本助手</h1><p>选择脚本 + 填写需求 → 生成新脚本 → 一键生成图片</p></div>
<div class="grid">
<div>
<div class="card"><h2>脚本文件</h2>
<div id="fileList">加载中...</div>
<button class="btn" onclick="loadFiles()">刷新</button></div>
<div class="card"><h2>客户需求</h2>
<textarea id="requirements" placeholder="产品介绍、目标用户、风格要求、时长、投放平台等"></textarea></div>
<div class="card"><h2>参考链接</h2>
<textarea id="urls" placeholder="每行一个参考链接"></textarea></div>
</div>
<div>
<div class="card"><h2>AI对话</h2>
<div id="chat" class="chat">你好！请选择脚本文件或填写需求开始。</div>
<input id="input" placeholder="请输入..." style="width:70%;background:#2C2C2E;border:none;border-radius:8px;padding:10px;color:#fff">
<button class="btn" onclick="send()">发送</button></div>
<div class="card"><h2>生成的脚本</h2>
<div id="script"></div>
<button class="btn" onclick="generate()">生成脚本</button>
<button class="btn" onclick="download()">导出Excel</button>
<button class="btn" onclick="toJimeng()">一键生成图片</button></div>
</div></div></div>
<script>
let currentScript=[];
function loadFiles(){
fetch('/files').then(r=>r.json()).then(d=>{
document.getElementById('fileList').innerHTML=d.files.map((f,i)=>'<div class="file-item" onclick="readFile('+i+')">'+f.name+'</div>').join('');
files=d.files;
});
}
function readFile(i){
fetch('/read',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name:files[i].name})})
.then(r=>r.json()).then(d=>{
currentScript=d.prompts||[];
document.getElementById('chat').innerHTML+='<div class="msg msg-ai">已加载 '+currentScript.length+' 个分镜</div>';
});
}
function send(){
let msg=document.getElementById('input').value;
if(!msg)return;
document.getElementById('chat').innerHTML+='<div class="msg msg-user">'+msg+'</div>';
document.getElementById('input').value='';
fetch('/chat',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({msg:msg,req:document.getElementById('requirements').value})})
.then(r=>r.json()).then(d=>{
document.getElementById('chat').innerHTML+='<div class="msg msg-ai">'+d.reply+'</div>';
});
}
function generate(){
let req=document.getElementById('requirements').value;
fetch('/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({req:req,scripts:currentScript})})
.then(r=>r.json()).then(d=>{
if(d.scripts){
currentScript=d.scripts;
document.getElementById('script').innerHTML=d.scripts.map((s,i)=>'<div>'+(i+1)+'. '+s+'</div>').join('');
}
});
}
function download(){
if(!currentScript.length)return alert('请先生成脚本');
location.href='/download?data='+encodeURIComponent(currentScript.join('|'));
}
function toJimeng(){
if(!currentScript.length)return alert('请先生成脚本');
fetch('/queue',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({scripts:currentScript})})
.then(r=>r.json()).then(d=>{
document.getElementById('chat').innerHTML+='<div class="msg msg-ai">已添加 '+currentScript.length+' 个任务到生成队列</div>';
});
}
loadFiles();
</script></body></html>
'''

@app.route('/')
def index(): return HTML

@app.route('/files')
def files():
    desktop = os.path.expanduser('~/Desktop')
    return {'files': [{'name':f} for f in os.listdir(desktop) if f.endswith('.xlsx')]}

@app.route('/read', methods=['POST'])
def read():
    try:
        wb = openpyxl.load_workbook(os.path.expanduser('~/Desktop/'+request.json['name']))
        ws = wb.active
        prompts = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] and isinstance(row[0], int) and row[1]:
                prompts.append(str(row[1]))
        return {'prompts': prompts}
    except Exception as e:
        return {'prompts': [], 'error': str(e)}

@app.route('/chat', methods=['POST'])
def chat():
    return {'reply': f"收到：{request.json.get('msg','')}。请填写客户需求后点击生成脚本。"}

@app.route('/generate', methods=['POST'])
def generate():
    req = request.json.get('req','')
    scripts = request.json.get('scripts',[])
    # 简单生成
    new_scripts = [f"展示产品：{req[:20]}..." for _ in range(10)]
    return {'scripts': new_scripts if not scripts else scripts}

@app.route('/download')
def download():
    data = request.args.get('data','').split('|')
    wb = openpyxl.Workbook(); ws = wb.active
    for i,s in enumerate(data,1):
        ws[f'A{i}']=i; ws[f'B{i}']=s
    wb.save('/tmp/out.xlsx')
    return send_file('/tmp/out.xlsx')

@app.route('/queue', methods=['POST'])
def queue():
    open('/tmp/jq.json','w').write(str(request.json.get('scripts',[])))
    return {'ok':True}

if __name__=='__main__':
    print("脚本助手:5002")
    app.run(host='0.0.0.0',port=5002,debug=False)
