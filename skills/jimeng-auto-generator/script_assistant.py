#!/usr/bin/env python3
"""
即梦AI脚本生成助手 - 专业版
AI辅助生成视频分镜脚本
"""
import os
import json
import time
import glob
from flask import Flask, render_template_string, request, jsonify, send_file
import openpyxl

app = Flask(__name__)
app.secret_key = 'script-assistant-v1'

# 全局状态
状态 = {
    'current_script': None,
    'chat_history': [],
    'requirements': '',
    'reference_urls': '',
    'generated_script': []
}

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AI脚本助手 - 专业版</title>
    <style>
        :root {
            --primary: #007AFF;
            --bg: #000000;
            --card: #1C1C1E;
            --card2: #2C2C2E;
            --text: #FFFFFF;
            --text2: #8E8E93;
            --green: #30D158;
            --orange: #FF9F0A;
        }
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, sans-serif;
            background: var(--bg);
            color: var(--text);
            min-height: 100vh;
        }
        .container { max-width: 1400px; margin: 0 auto; padding: 20px; }
        
        .header {
            text-align: center;
            padding: 30px;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            border-radius: 20px;
            margin-bottom: 20px;
        }
        .header h1 { font-size: 36px; margin-bottom: 10px; }
        .header p { color: var(--text2); font-size: 16px; }
        
        .grid { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
        .card {
            background: var(--card);
            border-radius: 16px;
            padding: 24px;
        }
        .card h2 { font-size: 20px; margin-bottom: 16px; display: flex; align-items: center; gap: 10px; }
        
        /* 文件列表 */
        .file-list { max-height: 300px; overflow-y: auto; }
        .file-item {
            padding: 12px 16px;
            background: var(--card2);
            border-radius: 10px;
            margin-bottom: 8px;
            cursor: pointer;
            transition: all 0.2s;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .file-item:hover { background: #3A3A3C; }
        .file-item.active { border: 2px solid var(--primary); }
        
        /* 输入框 */
        textarea, input[type="text"] {
            width: 100%;
            background: var(--card2);
            border: none;
            border-radius: 10px;
            padding: 14px;
            color: var(--text);
            font-size: 14px;
            margin-bottom: 12px;
        }
        textarea:focus, input:focus { outline: 2px solid var(--primary); }
        textarea { min-height: 120px; resize: vertical; }
        
        /* 按钮 */
        .btn {
            padding: 12px 24px;
            border-radius: 10px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            border: none;
            margin-right: 8px;
            margin-bottom: 8px;
        }
        .btn-primary { background: var(--primary); color: white; }
        .btn-primary:hover { background: #0077ED; }
        .btn-green { background: var(--green); color: black; }
        
        /* 聊天 */
        .chat-box {
            height: 400px;
            overflow-y: auto;
            background: var(--card2);
            border-radius: 12px;
            padding: 16px;
            margin-bottom: 16px;
        }
        .message {
            margin-bottom: 12px;
            padding: 12px 16px;
            border-radius: 12px;
            max-width: 80%;
        }
        .message.ai { background: var(--card); align-self: flex-start; }
        .message.user { background: var(--primary); margin-left: auto; }
        
        /* 生成的脚本 */
        .script-preview {
            max-height: 500px;
            overflow-y: auto;
        }
        .script-item {
            padding: 12px;
            background: var(--card2);
            border-radius: 8px;
            margin-bottom: 8px;
            font-size: 14px;
        }
        .script-item .num {
            color: var(--primary);
            font-weight: bold;
            margin-right: 8px;
        }
        
        /* 标签页 */
        .tabs { display: flex; gap: 10px; margin-bottom: 16px; }
        .tab {
            padding: 8px 16px;
            background: var(--card2);
            border-radius: 8px;
            cursor: pointer;
        }
        .tab.active { background: var(--primary); }
        
        @media (max-width: 768px) {
            .grid { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎬 AI脚本助手</h1>
            <p>专业的视频分镜脚本生成工具 | 结合客户需求生成最佳脚本</p>
        </div>
        
        <div class="grid">
            <!-- 左侧：文件+需求 -->
            <div>
                <div class="card" style="margin-bottom: 20px;">
                    <h2>📁 现有脚本库</h2>
                    <div class="file-list" id="fileList">
                        <div style="color:var(--text2);text-align:center;padding:20px;">加载中...</div>
                    </div>
                    <button class="btn btn-primary" onclick="loadFiles()" style="margin-top:12px;">🔄 刷新文件列表</button>
                </div>
                
                <div class="card">
                    <h2>📋 客户需求</h2>
                    <textarea id="requirements" placeholder="请详细描述客户需求，例如：
- 产品是什么（功能、特点）
- 目标用户群体
- 期望的视频风格
- 重点展示的卖点
- 时长要求
- 投放平台"></textarea>
                </div>
                
                <div class="card" style="margin-top:20px;">
                    <h2>🔗 参考链接</h2>
                    <textarea id="referenceUrls" placeholder="请粘贴参考视频/图片链接，每行一个"></textarea>
                </div>
            </div>
            
            <!-- 右侧：AI对话+生成 -->
            <div>
                <div class="card">
                    <h2>💬 与AI助手对话</h2>
                    <div class="chat-box" id="chatBox">
                        <div class="message ai">你好！我是AI脚本助手。请先：<br>1. 选择或上传脚本文件<br>2. 填写客户需求<br>3. 点击生成脚本<br><br>我可以帮你：<br>- 分析现有脚本<br>- 根据需求生成新脚本<br>- 优化脚本细节<br>- 添加专业建议</div>
                    </div>
                    <div style="display:flex;gap:8px;">
                        <input type="text" id="userInput" placeholder="请输入您的问题..." style="flex:1;margin:0;">
                        <button class="btn btn-primary" onclick="sendMessage()">发送</button>
                    </div>
                </div>
                
                <div class="card" style="margin-top:20px;">
                    <h2>📝 生成的脚本</h2>
                    <div class="tabs">
                        <div class="tab active" onclick="showTab('preview')">脚本预览</div>
                        <div class="tab" onclick="showTab('edit')">编辑脚本</div>
                    </div>
                    <div id="scriptPreview" class="script-preview">
                        <div style="color:var(--text2);text-align:center;padding:40px;">请先生成脚本</div>
                    </div>
                    <div id="scriptEdit" style="display:none;">
                        <textarea id="scriptEditArea" style="min-height:300px;"></textarea>
                    </div>
                    <div style="margin-top:16px;">
                        <button class="btn btn-primary" onclick="generateScript()">✨ 生成脚本</button>
                        <button class="btn btn-green" onclick="downloadExcel()">📥 导出Excel</button>
                        <button class="btn btn-primary" onclick="sendToJimeng()">🎨 一键生成图片</button>
                    </div>
                </div>
            </div>
        </div>
    </div>
    
    <script>
        let files = [];
        let currentFile = null;
        let chatHistory = [];
        
        // 加载文件列表
        async function loadFiles() {
            const res = await fetch('/list_files');
            const data = await res.json();
            files = data.files;
            
            const list = document.getElementById('fileList');
            if (files.length === 0) {
                list.innerHTML = '<div style="color:var(--text2);text-align:center;">桌面上没有xlsx文件</div>';
                return;
            }
            
            list.innerHTML = '<div onclick="selectFile(0)">点击加载文件</div>';
        }
        
        async function selectFile(idx) {
        }                    <span style="color:var(--text2);font-size:12px;">${f.size}</span>
                </div>
            """).join('');
        }
        
        // 选择文件
        async function selectFile(index) {
            currentFile = files[index];
            document.querySelectorAll('.file-item').forEach((el, i) => {
                el.classList.toggle('active', i === index);
            });
            
            // 读取文件内容
            const res = await fetch('/read_file', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({filename: currentFile.name})
            });
            const data = await res.json();
            
            if (data.success) {
                addMessage('user', '我选择了文件：' + currentFile.name);
                addMessage('ai', '已加载文件，包含 ' + data.prompts.length + ' 个分镜。\\n\\n' + 
                    '请填写客户需求，然后点击"生成脚本"或直接问我问题。');
            }
        }
        
        // 发送消息
        async function sendMessage() {
            const input = document.getElementById('userInput');
            const msg = input.value.trim();
            if (!msg) return;
            
            addMessage('user', msg);
            input.value = '';
            
            const res = await fetch('/chat', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    message: msg,
                    context: {
                        requirements: document.getElementById('requirements').value,
                        referenceUrls: document.getElementById('referenceUrls').value,
                        currentFile: currentFile?.name
                    }
                })
            });
            const data = await res.json();
            addMessage('ai', data.response);
        }
        
        // 添加消息
        function addMessage(role, content) {
            const box = document.getElementById('chatBox');
            box.innerHTML += '<div class="message ' + role + '">' + content.replace(/\\n/g, '<br>') + '</div>';
            box.scrollTop = box.scrollHeight;
        }
        
        // 生成脚本
        async function generateScript() {
            const requirements = document.getElementById('requirements').value;
            const referenceUrls = document.getElementById('referenceUrls').value;
            
            if (!requirements) {
                alert('请先填写客户需求');
                return;
            }
            
            addMessage('user', '请根据需求生成脚本');
            addMessage('ai', '正在根据需求生成专业脚本，请稍候...');
            
            const res = await fetch('/generate_script', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    requirements: requirements,
                    referenceUrls: referenceUrls,
                    currentFile: currentFile?.name
                })
            });
            const data = await res.json();
            
            if (data.success) {
                // 显示生成的脚本
                const preview = document.getElementById('scriptPreview');
                preview.innerHTML = data.script.map((s, i) => 
                    '<div class="script-item"><span class="num">' + (i+1) + '</span>' + s + '</div>'
                ).join('');
                
                document.getElementById('scriptEditArea').value = data.script.join('\\n');
                addMessage('ai', '✅ 脚本生成完成！共 ' + data.script.length + ' 个分镜。\\n\\n你可以：\\n- 点击"编辑脚本"修改内容\\n- 点击"导出Excel"下载文件\\n- 点击"一键生成图片"到即梦AI生成图片');
            } else {
                addMessage('ai', '生成失败：' + data.error);
            }
        }
        
        // 导出Excel
        async function downloadExcel() {
            const script = document.getElementById('scriptEditArea').value.split('\\n').filter(s => s.trim());
            if (script.length === 0) {
                alert('请先生成脚本');
                return;
            }
            
            window.location.href = '/download_excel?script=' + encodeURIComponent(script.join('|'));
        }
        
        // 一键生成图片
        async function sendToJimeng() {
            const script = document.getElementById('scriptEditArea').value.split('\\n').filter(s => s.trim());
            if (script.length === 0) {
                alert('请先生成脚本');
                return;
            }
            
            const res = await fetch('/queue_to_jimeng', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({script: script})
            });
            const data = await res.json();
            
            if (data.success) {
                addMessage('ai', '✅ 已将 ' + script.length + ' 个任务添加到生成队列！\\n\\n请到图片生成页面查看进度。');
            } else {
                addMessage('ai', '添加到队列失败：' + data.error);
            }
        }
        
        // 显示标签页
        function showTab(tab) {
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            event.target.classList.add('active');
            
            document.getElementById('scriptPreview').style.display = tab === 'preview' ? 'block' : 'none';
            document.getElementById('scriptEdit').style.display = tab === 'edit' ? 'block' : 'none';
        }
        
        // 初始化
        loadFiles();
        
        // 回车发送
        document.getElementById('userInput').addEventListener('keypress', e => {
            if (e.key === 'Enter') sendMessage();
        });
    </script>
</body>
</html>
"""

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/list_files')
def list_files():
    """列出桌面上的xlsx文件"""
    desktop = os.path.expanduser('~/Desktop')
    files = []
    for f in glob.glob(os.path.join(desktop, '*.xlsx')):
        size = os.path.getsize(f) / 1024 / 1024
        files.append({
            'name': os.path.basename(f),
            'path': f,
            'size': f'{size:.1f}MB'
        })
    return jsonify({'files': files})

@app.route('/read_file', methods=['POST'])
def read_file():
    """读取xlsx文件内容"""
    data = request.json
    filename = data.get('filename')
    
    filepath = os.path.expanduser(f'~/Desktop/{filename}')
    
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        prompts = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] and isinstance(row[0], int) and row[1]:
                prompts.append(str(row[1]))
        
        return jsonify({'success': True, 'prompts': prompts, 'count': len(prompts)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/chat', methods=['POST'])
def chat():
    """AI对话"""
    data = request.json
    message = data.get('message', '')
    context = data.get('context', {})
    
    # 简单的响应
    response = f"收到你的消息：{message}\n\n"
    
    if context.get('requirements'):
        response += f"客户需求：{context['requirements'][:100]}...\n"
    
    response += "\n请问还有什么需要帮助的？"
    
    return jsonify({'response': response})

@app.route('/generate_script', methods=['POST'])
def generate_script():
    """生成脚本"""
    data = request.json
    requirements = data.get('requirements', '')
    referenceUrls = data.get('referenceUrls', '')
    
    # 生成详细的脚本
    script = []
    
    # 解析需求关键词
    product_keywords = []
    for w in ['产品', '按摩器', '净化器', '狗狗', '猫', '衣服', '道具']:
        if w in requirements:
            product_keywords.append(w)
    
    if not product_keywords:
        product_keywords = ['产品']
    
    # 生成10个分镜
    for i in range(10):
        if i == 0:
            script.append(f"展示{product_keywords[0]}整体外观，{requirements[:30] if requirements else '高清专业拍摄'}")
        elif i == 3:
            script.append(f"展示{product_keywords[0]}使用场景，{requirements[:30] if requirements else '温馨氛围'}")
        elif i == 6:
            script.append(f"详细展示{product_keywords[0]}功能特点")
        elif i == 9:
            script.append(f"总结{product_keywords[0]}优势，呼吁购买")
        else:
            script.append(f"镜头{i+1}: 继续展示{product_keywords[0]}，{requirements[:20] if requirements else '专业摄影'}")
    
    return jsonify({'success': True, 'script': script})

@app.route('/download_excel')
def download_excel():
    """下载Excel"""
    script_str = request.args.get('script', '')
    script = script_str.split('|')
    
    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "视频分镜"
    
    # 写入数据
    ws['A1'] = '镜号'
    ws['B1'] = '分镜描述'
    
    for i, s in enumerate(script, 1):
        ws[f'A{i+1}'] = i
        ws[f'B{i+1}'] = s
    
    # 保存
    filepath = '/tmp/generated_script.xlsx'
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name='视频分镜脚本.xlsx')

@app.route('/queue_to_jimeng', methods=['POST'])
def queue_to_jimeng():
    """将脚本添加到即梦生成队列"""
    data = request.json
    script = data.get('script', [])
    
    # 保存到队列文件
    queue_file = '/tmp/jimeng_queue.json'
    with open(queue_file, 'w', encoding='utf-8') as f:
        json.dump(script, f, ensure_ascii=False)
    
    return jsonify({'success': True, 'count': len(script)})

if __name__ == '__main__':
    print("🎬 AI脚本助手启动...")
    print("📍 访问地址: http://localhost:5002")
    app.run(host='0.0.0.0', port=5002, debug=False)
